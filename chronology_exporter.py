#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel 导出：多年份多Sheet，季度分组，样式美化"""

import os
import logging
from typing import List, Dict, Any
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from config import (
    OUTPUT_DIR, SEASONS, SEASON_ORDER, SEASON_COLORS,
    SEASON_HEADER_COLORS, FIELD_NAMES, WEEKDAY_CN,
)

logger = logging.getLogger(__name__)

COLUMNS = [
    ("name_jp", 30),
    ("name_cn", 30),
    ("category", 8),
    ("episodes", 6),
    ("air_date", 12),
    ("air_weekday", 8),
    ("studio", 20),
    ("score", 6),
    ("watch_status", 10),
]

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGN = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

SEASON_TITLE_FONT = Font(name="微软雅黑", bold=True, size=13, color="FFFFFF")

HIGH_SCORE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


class ChronologyExporter:
    def __init__(self, output_dir: str = OUTPUT_DIR):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def export(self, anime_list: List[Dict[str, Any]],
               filename: str = None) -> str:
        if not anime_list:
            logger.warning("数据为空，不导出")
            return ""

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d")
            filename = f"番剧年表_{timestamp}.xlsx"

        filepath = os.path.join(self.output_dir, filename)
        wb = Workbook()

        years = self._group_by_year(anime_list)
        self._write_overview(wb, anime_list, years)

        for year in sorted(years.keys()):
            ws = wb.create_sheet(title=f"{year}年")
            self._write_year_sheet(ws, year, years[year])

        ws_all = wb.create_sheet(title="全部数据")
        self._write_flat_sheet(ws_all, anime_list)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        try:
            wb.save(filepath)
            logger.info(f"Excel 已导出: {filepath} ({len(anime_list)} 条)")
        except Exception as e:
            logger.error(f"保存 Excel 失败: {e}")
            return ""

        return filepath

    def _group_by_year(self, anime_list: List[Dict[str, Any]]) -> Dict[int, List]:
        years = {}
        for anime in anime_list:
            year = anime.get("year", 0)
            if year:
                years.setdefault(year, []).append(anime)
        return years

    def _write_overview(self, wb: Workbook, anime_list: List[Dict[str, Any]],
                        years: Dict[int, List]):
        ws = wb.active
        ws.title = "总览"

        ws.merge_cells("A1:F1")
        cell = ws["A1"]
        cell.value = "番剧年表 - 总览"
        cell.font = Font(name="微软雅黑", bold=True, size=16)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        headers = ["年份", "一月", "四月", "七月", "十月", "合计"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = HEADER_ALIGN
            c.border = THIN_BORDER

        row = 4
        for year in sorted(years.keys()):
            year_anime = years[year]
            season_counts = {"一月": 0, "四月": 0, "七月": 0, "十月": 0}
            unseasoned = 0
            for a in year_anime:
                s = a.get("season", "")
                if s in season_counts:
                    season_counts[s] += 1
                else:
                    unseasoned += 1

            ws.cell(row=row, column=1, value=year).alignment = CENTER_ALIGN
            for si, season in enumerate(SEASON_ORDER):
                ws.cell(row=row, column=si + 2, value=season_counts[season]).alignment = CENTER_ALIGN
            ws.cell(row=row, column=6, value=len(year_anime)).alignment = CENTER_ALIGN

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).font = DATA_FONT

            row += 1

        for col, w in enumerate([8, 8, 8, 8, 8, 8], 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.freeze_panes = "A4"

    def _write_year_sheet(self, ws, year: int, anime_list: List[Dict[str, Any]]):
        season_groups = {}
        for s in SEASON_ORDER:
            season_groups[s] = [a for a in anime_list if a.get("season") == s]
        unseasoned = [a for a in anime_list if a.get("season") not in SEASON_ORDER]

        current_row = 1

        for season in SEASON_ORDER:
            season_anime = season_groups[season]
            if not season_anime:
                continue

            season_info = SEASONS[season]
            header_color = SEASON_HEADER_COLORS.get(season, "366092")
            bg_color = SEASON_COLORS.get(season, "FFFFFF")

            title = f"{year}年 {season}番（{season_info['label']}）"
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=len(COLUMNS))
            cell = ws.cell(row=current_row, column=1, value=title)
            cell.font = SEASON_TITLE_FONT
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[current_row].height = 30
            current_row += 1

            for col, (field, width) in enumerate(COLUMNS, 1):
                c = ws.cell(row=current_row, column=col, value=FIELD_NAMES.get(field, field))
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = HEADER_ALIGN
                c.border = THIN_BORDER
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            for idx, anime in enumerate(season_anime):
                row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid") if idx % 2 == 0 else None

                for col, (field, _) in enumerate(COLUMNS, 1):
                    val = anime.get(field, "")
                    c = ws.cell(row=current_row, column=col, value=val)
                    c.font = DATA_FONT
                    c.border = THIN_BORDER
                    c.alignment = DATA_ALIGN

                    if field in ("category", "episodes", "air_weekday", "score", "watch_status"):
                        c.alignment = CENTER_ALIGN

                    if row_fill:
                        c.fill = row_fill

                score = anime.get("score", 0)
                if score and score >= 8.0:
                    for col in range(1, len(COLUMNS) + 1):
                        ws.cell(row=current_row, column=col).fill = HIGH_SCORE_FILL

                ws.row_dimensions[current_row].height = 20
                current_row += 1

            current_row += 1

        for col, (_, width) in enumerate(COLUMNS, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"

    def _write_flat_sheet(self, ws, anime_list: List[Dict[str, Any]]):
        all_columns = [
            ("year", 8), ("season", 8),
            ("name_jp", 30), ("name_cn", 30),
            ("category", 8), ("episodes", 6),
            ("air_date", 12), ("air_weekday", 8),
            ("studio", 20), ("score", 6), ("watch_status", 10),
        ]

        for col, (field, _) in enumerate(all_columns, 1):
            c = ws.cell(row=1, column=col, value=FIELD_NAMES.get(field, field))
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = HEADER_ALIGN
            c.border = THIN_BORDER

        for row_idx, anime in enumerate(anime_list, 2):
            for col, (field, _) in enumerate(all_columns, 1):
                val = anime.get(field, "")
                c = ws.cell(row=row_idx, column=col, value=val)
                c.font = DATA_FONT
                c.border = THIN_BORDER
                c.alignment = DATA_ALIGN
                if field in ("year", "season", "category", "episodes", "air_weekday", "score", "watch_status"):
                    c.alignment = CENTER_ALIGN

        for col, (_, width) in enumerate(all_columns, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def export_by_year(self, anime_list: List[Dict[str, Any]],
                       output_dir: str = None) -> List[str]:
        if not anime_list:
            return []

        out_dir = output_dir or os.path.join(self.output_dir, "按年份")
        os.makedirs(out_dir, exist_ok=True)

        years = self._group_by_year(anime_list)
        generated = []

        for year in sorted(years.keys()):
            filename = f"番剧年表_{year}年.xlsx"
            filepath = os.path.join(out_dir, filename)

            wb = Workbook()
            ws = wb.active
            ws.title = f"{year}年"
            self._write_year_sheet(ws, year, years[year])

            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Sheet"]

            try:
                wb.save(filepath)
                generated.append(filepath)
                logger.info(f"已导出 {year}年: {filepath}")
            except Exception as e:
                logger.error(f"导出 {year}年失败: {e}")

        return generated
